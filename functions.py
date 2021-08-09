import openpyxl
import string
import random

from classes import *


# Constants
Team_number = 10
Teams = [
"Ac Chiavo V.",
"Ac M.R. King",
"AC Panz.-ott.",
"Al-Akhazam",
"As Stronzi",
"I Miserabili",
"Juve Scabbia",
"La La Lecco",
"Post Malore",
"Sampidoria"
]

# Setup

def orderpick(TeamList):
    TeamList = random.shuffle(TeamList)

# Data extraction

# DISCLAIMER: Fantacalcio's data must be ordered by "Quotazione attuale".

def extractor(sheet):
    l = []
    for i in range(len(sheet['A'])-2):
            l.append(Player(sheet.cell(row=i+3, column=1).value,str(sheet.cell(row=i+3, column=3).value),str(sheet.cell(row=i+3, column=4).value),str(sheet.cell(row=i+3, column=5).value),str(sheet.cell(row=i+3, column=6).value)))
    return l


def get_players_list(inputFile):

    goalkeepers = []
    defenders = []
    midfielders = []
    forwards = []

    doc = openpyxl.load_workbook(inputFile)
    sheet_gk = doc.get_sheet_by_name('Portieri')
    sheet_def = doc.get_sheet_by_name('Difensori')
    sheet_mid = doc.get_sheet_by_name('Centrocampisti')
    sheet_fw = doc.get_sheet_by_name('Attaccanti')


    goalkeepers = extractor(sheet_gk)
    defenders = extractor(sheet_def)
    midfielders = extractor(sheet_mid)
    forwards = extractor(sheet_fw)

    return goalkeepers, defenders, midfielders, forwards

# Math

def roasterValue(TeamList):
    list = []
    for i in len(TeamList):
        list.append(TeamList[i][3])

    value = sum(list)

    return value



# Team constructor

def get_and_remove_elite(teamlist,list):
    elite = list[:5]
    pick = random.choice(elite)
    teamlist.append(pick)
    list.remove(pick)
    return teamlist, list

def get_and_remove(teamlist,list):
    pick = random.choice(list)
    teamlist.append(pick)
    list.remove(pick)
    return teamlist, list

def get_and_remove_elite_gk(teamlist, list):
    elite = list[:5]
    pick = random.choice(elite)
    team = pick.team

    teamlist.add_gk(pick)
    list.remove(pick)

    for sublist in list:
        if sublist.team == team:
            teamlist.add_gk(sublist)
            list.remove(sublist)
            break
    for sublist in list:
        if sublist.team == team:
            #teamlist.append(sublist)
            list.remove(sublist) #Rumuovo il terzo portiere eventualmente
            break
    return teamlist, list

def get_and_remove_elite_def(teamlist, list):
    elite = list[:5]
    pick = random.choice(elite)

    teamlist.add_def(pick)
    list.remove(pick)

    return teamlist, list

def get_and_remove_elite_mid(teamlist, list):
    elite = list[:5]
    pick = random.choice(elite)

    teamlist.add_mid(pick)
    list.remove(pick)

    return teamlist, list

def get_and_remove_elite_fw(teamlist, list):
    elite = list[:5]
    pick = random.choice(elite)

    teamlist.add_fw(pick)
    list.remove(pick)

    return teamlist, list
